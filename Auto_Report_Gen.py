# Import necessary libraries
import pandas as pd  # For data manipulation
import numpy as np  # For numerical operations
import warnings  # To handle warnings
import sys  # For system-related functions
import os  # For operating system tasks
import platform  # For platform information
import time  # For time-related functions
from colorama import Fore  # For colored text in the terminal
from openpyxl import Workbook  # For working with Excel files
from openpyxl.utils import get_column_letter  # To convert column numbers to letters
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill  # For styling Excel cells
from datetime import datetime  # For date and time handling

# Ignore warnings to keep output clean
warnings.filterwarnings('ignore')

def clear_console():
    """
    Clear the console screen.

    This function checks the operating system and executes the appropriate command
    to clear the console. It uses 'cls' for Windows and 'clear' for Unix/Linux/Mac.

    Returns:
        None
    """
    if platform.system() == "Windows":
        os.system('cls')  # For Windows
    else:
        os.system('clear')  # For Unix/Linux/Mac

def display_banner():
    """
    Display a decorative banner in the console.

    This function prints a stylized banner that includes information about the 
    Auto Correlation and Cpk Report Generator, along with the author's name, 
    email, and GitHub link. The banner is displayed in cyan color and is followed 
    by a brief pause.

    Returns:
        None
    """
    banner = '''
┏┓       ┏┓        ┓   •           ┓  ┏┓  ┓   ┳┓           ┏┓               
┣┫┓┏╋┏┓  ┃ ┏┓┏┓┏┓┏┓┃┏┓╋┓┏┓┏┓  ┏┓┏┓┏┫  ┃ ┏┓┃┏  ┣┫┏┓┏┓┏┓┏┓╋  ┃┓┏┓┏┓┏┓┏┓┏┓╋┏┓┏┓
┛┗┗┻┗┗┛  ┗┛┗┛┛ ┛ ┗ ┗┗┻┗┗┗┛┛┗  ┗┻┛┗┗┻  ┗┛┣┛┛┗  ┛┗┗ ┣┛┗┛┛ ┗  ┗┛┗ ┛┗┗ ┛ ┗┻┗┗┛┛ 
                                        ┛         ┛                                                                                                              
Auto Correlation and Cpk Report Generator
Coded by Mohamad Haikal bin Mohamad Nazari
Email: mohamadhaikal.mohamadnazari@tessolve.com
Github: https://github.com/haikal5e
    '''
    print(f"{Fore.CYAN}{banner}{Fore.RESET}")
    time.sleep(1)

def thank_you():
    """
    Display a thank you message in the console.

    This function prints a stylized thank you message in green color. The message 
    is displayed in a decorative format and is followed by a brief pause.

    Returns:
        None
    """
    thank = '''
┏┳┓┓     ┓   ┓┏    
 ┃ ┣┓┏┓┏┓┃┏  ┗┫┏┓┓┏
 ┻ ┛┗┗┻┛┗┛┗  ┗┛┗┛┗┻
                                                                                                              
    '''
    
    print(f"{Fore.GREEN}{thank}{Fore.RESET}")
    time.sleep(1)

def get_product_information():
    """
    Collect and manage product information with user verification.

    This function prompts the user to enter various product details, including 
    the test card name, part name, package, lead count, and description. It 
    ensures that the user cannot leave any fields empty and allows for 
    reviewing and modifying the entered information before final confirmation.

    Returns:
        dict: A dictionary containing the collected product information.
    """
    
    # Initialize product information dictionary
    product_info = {
        'Test Card Name': '',
        'Part Name': '',
        'Package': '',
        'Lead Count': '',
        'Description': ''
    }
    
    # Collect initial product information
    print("Please enter the following product information:")
    
    for key in product_info.keys():
        while True:
            value = input(f"Enter {key}: ").strip()
            if value:
                product_info[key] = value
                break
            else:
                print()
                print(f"{key} cannot be empty. Please try again.")

    print()  # Retained for separation
    
    # Review and confirmation loop
    while True:
        clear_console()
        print("Current Product Information:")
        for key, value in product_info.items():
            print(f"{key}: {value}")
        
        # Confirmation and change option
        confirm = input("\nAre you satisfied with the current information? (yes/no): ").strip().lower()
        
        if confirm == 'yes':
            print()  # Retained for separation
            break
        elif confirm == 'no':
            print()  # Retained for separation
            # Allow specific field changes
            while True:
                clear_console()
                print("Current Product Information:")
                for key, value in product_info.items():
                    print(f"{key}: {value}")
                print()  # Retained for separation
                
                # Field selection for modification
                print("Which field would you like to change?")
                for i, key in enumerate(product_info.keys(), 1):
                    print(f"{i}. {key}")
                print("0. Go back to confirmation")
                print()  # Retained for separation
                
                try:
                    choice = int(input("Enter the number of the field you want to modify: "))
                    
                    # print()  # Retained for separation
                    
                    if choice == 0:
                        break
                    elif 1 <= choice <= len(product_info):
                        field = list(product_info.keys())[choice - 1]
                        print()  # Retained for separation
                        new_value = input(f"Enter new {field}: ").strip()
                        
                        # Validate input
                        if new_value:
                            product_info[field] = new_value
                            print()  # Retained for separation
                            print(f"{field} updated successfully!")
                        else:
                            print()  # Retained for separation
                            print("Invalid input. Field cannot be empty.")
                        
                        input("\nPress Enter to continue...\n")
                    else:
                        print()  # Retained for separation
                        print("Invalid choice. Please try again.")
                        input("\nPress Enter to continue...\n")
                
                except ValueError:
                    print()  # Retained for separation
                    print("Please enter a valid number.")
                    input("\nPress Enter to continue...\n")
        else:
            print()  # Retained for separation
            print("Invalid input. Please enter 'yes' or 'no'.")
            input("\nPress Enter to continue...\n")
    
    return product_info

def get_setup_information():
    """
    Collect and manage setup information with user verification.

    This function prompts the user to enter various setup details, including 
    the tester ID, reference board, new board ID, and test program. It ensures 
    that the user cannot leave any fields empty and allows for reviewing and 
    modifying the entered information before final confirmation.

    Returns:
        dict: A dictionary containing the collected setup information.
    """
    clear_console()
    
    # Initialize setup information dictionary
    setup_info = {
        'Tester ID': '',
        'Reference Board': '',
        'New Board ID': '',
        'Test Program': ''
    }
    
    # Collect initial setup information
    print("Please enter the following setup information:")
    
    for key in setup_info.keys():
        while True:
            value = input(f"Enter {key}: ").strip()
            if value:
                setup_info[key] = value
                break
            else:
                print()
                print(f"{key} cannot be empty. Please try again.")

    print()  # Retained for separation
    
    # Review and confirmation loop
    while True:
        clear_console()
        print("Current Setup Information:")
        for key, value in setup_info.items():
            print(f"{key}: {value}")
        
        # Confirmation and change option
        confirm = input("\nAre you satisfied with the current setup information? (yes/no): ").strip().lower()
        
        if confirm == 'yes':
            print()  # Retained for separation
            break
        elif confirm == 'no':
            print()  # Retained for separation
            # Allow specific field changes
            while True:
                clear_console()
                print("Current Setup Information:")
                for key, value in setup_info.items():
                    print(f"{key}: {value}")
                print()  # Retained for separation
                
                # Field selection for modification
                print("Which field would you like to change?")
                for i, key in enumerate(setup_info.keys(), 1):
                    print(f"{i}. {key}")
                print("0. Go back to confirmation")
                print()  # Retained for separation
                
                try:
                    choice = int(input("Enter the number of the field you want to modify: "))
                    
                    if choice == 0:
                        break
                    elif 1 <= choice <= len(setup_info):
                        print()  # Retained for separation
                        field = list(setup_info.keys())[choice - 1]
                        new_value = input(f"Enter new {field}: ").strip()
                        
                        # Validate input
                        if new_value:
                            setup_info[field] = new_value
                            print()
                            print(f"{field} updated successfully!")
                        else:
                            print("\nInvalid input. Field cannot be empty.")
                        
                        input("\nPress Enter to continue...\n")
                    else:
                        print("\nInvalid choice. Please try again.")
                        input("\nPress Enter to continue...\n")
                
                except ValueError:
                    print("\nPlease enter a valid number.")
                    input("\nPress Enter to continue...\n")
        else:
            print("\nInvalid input. Please enter 'yes' or 'no'.")
            input("\nPress Enter to continue...\n")
    
    return setup_info

def get_data():
    """
    Gets input from the user about boards, units, and RB units, and returns file path lists.

    This function prompts the user to enter the number of new boards and units tested, 
    ensuring the values are between 1 and 9. It collects file paths for each board and 
    unit, as well as a limit file path, and allows the user to review and modify the 
    entered paths before final confirmation.

    Returns:
        tuple: A tuple containing:
            - list: A list of lists with file paths for each board and unit.
            - list: A list of file paths for RB units.
            - int: The number of boards.
            - int: The number of units.
            - str: The file path for the limit file.
    """
    
    clear_console()  # Clear console at the start of the function
    
    while True:        
        try:
            num_boards = int(input("Enter the number of New Boards (NB) (between 1 and 9): "))
            if 1 <= num_boards <= 9:  # Check if the number is between 1 and 9
                break
            else:
                clear_console()
                print("Please enter a number between 1 and 9!")
                sys.stdout.flush()                
        except ValueError:
            clear_console()
            print("Invalid input. Please enter a valid number!")
            sys.stdout.flush()

    while True:
        try:
            num_units = int(input("Enter the number of units tested (between 1 and 9): "))
            if 1 <= num_units <= 9:  # Check if the number is between 1 and 9
                break
            else:
                print("Please enter a number between 1 and 9!")
                sys.stdout.flush()
        except ValueError:
            print("Invalid input. Please enter a valid number!")
            sys.stdout.flush()

    # Confirmation step
    while True:
        clear_console()
        confirmation = input(f"You entered {num_boards} boards and {num_units} units. Are you sure these amounts are correct? (yes/no): ").strip().lower()
        if confirmation == 'yes':
            break
        elif confirmation == 'no':
            clear_console()
            print("Let's enter the amounts again.")
            time.sleep(1.5)
            sys.stdout.flush()

            # Reset the input for number of boards and units
            while True:
                clear_console()
                try:
                    num_boards = int(input("Enter the number of New Boards (NB) (between 1 and 9): "))
                    if 1 <= num_boards <= 9:  # Check if the number is between 1 and 9
                        break
                    else:
                        print("Please enter a number between 1 and 9!")
                        sys.stdout.flush()
                except ValueError:
                    print("Invalid input. Please enter a valid number!")
                    sys.stdout.flush()

            while True:
                try:
                    num_units = int(input("Enter the number of units tested (between 1 and 9): "))
                    if 1 <= num_units <= 9:  # Check if the number is between 1 and 9
                        break
                    else:
                        print("Please enter a number between 1 and 9!")
                        sys.stdout.flush()
                except ValueError:
                    print("Invalid input. Please enter a valid number!")
                    sys.stdout.flush()
            continue  # Go back to the confirmation step after resetting the values
        else:
            clear_console()
            print("Invalid input. Please enter 'yes' or 'no'.")
            time.sleep(1.5)
            sys.stdout.flush()

    print()  # Necessary for spacing

    board_file_paths = []
    for board_num in range(num_boards):  # Range starts from 0
        board_paths = []
        for unit_num in range(num_units):  # Range starts from 0
            while True:
                file_path = input(f"Enter the file path for NB{board_num + 1} U{unit_num + 1} (CSV file): ")
                if file_path.lower().endswith(".csv"):
                    board_paths.append(file_path)
                    break
                else:
                    print() # Retained for separation
                    print("Invalid input. Please enter a file path that ends with '.csv'.")
                    sys.stdout.flush()
        print() # Retained for separation
        board_file_paths.append(board_paths)

    print()  # Necessary for spacing

    rb_file_paths = []
    for unit_num in range(num_units):  # Range starts from 0
        while True:
            file_path = input(f"Enter the file path for RB U{unit_num + 1} (CSV file): ")
            if file_path.lower().endswith (".csv"):
                rb_file_paths.append(file_path)
                break
            else:
                print() # Retained for separation
                print("Invalid input. Please enter a file path that ends with '.csv'.")
                sys.stdout.flush()
    
    print()  # Necessary for spacing
    
    while True:
        limit_file = input("Enter the file path for the limit file (CSV file): ")
        if limit_file.lower().endswith(".csv"):
            break
        else:
            print() # Retained for separation
            print("Invalid input. Please enter a file path that ends with '.csv'.")
            sys.stdout.flush()

    # Review and change file paths
    while True:
        clear_console()
        print("Current file paths:\n")
        for i, board_paths in enumerate(board_file_paths):
            for j, path in enumerate(board_paths):
                print(f"NB{(i + 1)} U{(j + 1)}: {path}")
        
        for j, path in enumerate(rb_file_paths):
            print(f"RB U{(j + 1)}: {path}")

        print(f"Limit file: {limit_file}")

        change = input("\nAre you satisfied with current file paths? (yes/no): ").strip().lower()
        if change == 'yes':
            break
        elif change == 'no':
            print() # Retained for separation
            while True:
                path_type = input("Which path do you want to change? (NB/RB/Limit): ").strip().lower()
                if path_type == 'nb':
                    print() # Retained for separation
                    # Validate board number
                    while True:
                        try:
                            board_index = int(input(f"Enter the new board number (1 to {num_boards}): ")) - 1
                            if 0 <= board_index < num_boards:
                                break
                            else:
                                print() # Retained for separation
                                print("Input out of range! Please enter a valid board number.")
                                sys.stdout.flush()
                        except ValueError:
                            print() # Retained for separation
                            print("Invalid input. Please enter a number.")
                            sys.stdout.flush()
                    
                    print() # Retained for separation
                    # Validate unit number
                    while True:
                        try:
                            unit_index = int(input(f"Enter the unit number (1 to {num_units}): ")) - 1
                            if 0 <= unit_index < num_units:
                                break
                            else:
                                print() # Retained for separation
                                print("Input out of range! Please enter a valid unit number.")
                                sys.stdout.flush()
                        except ValueError:
                            print() # Retained for separation
                            print("Invalid input. Please enter a number.")
                            sys.stdout.flush()
                    
                    print() # Retained for separation
                    # Get new file path
                    while True:
                        new_path = input(f"Enter the new file path for NB{board_index + 1} U{unit_index + 1} (CSV file): ")
                        if new_path.lower().endswith(".csv"):
                            board_file_paths[board_index][unit_index] = new_path
                            break
                        else:
                            print() # Retained for separation
                            print("Invalid path. It must end with '.csv'.")
                            sys.stdout.flush()

                elif path_type == 'rb':
                    print() # Retained for separation
                    # Validate unit number for RB
                    while True:
                        try:
                            unit_index = int(input(f"Enter the unit number (1 to {num_units}): ")) - 1
                            if 0 <= unit_index < num_units:
                                break
                            else:
                                print() # Retained for separation
                                print("Input out of range! Please enter a valid unit number.")
                                sys.stdout.flush()
                        except ValueError:
                            print()
                            print("Invalid input. Please enter a number.")
                            sys.stdout.flush()
                    
                    print() # Retained for separation
                    # Get new file path
                    while True:
                        new_path = input(f"Enter the new file path for RB U{unit_index + 1} (CSV file): ")
                        if new_path.lower().endswith(".csv"):
                            rb_file_paths[unit_index] = new_path
                            break
                        else:
                            print() # Retained for separation
                            print("Invalid path. It must end with '.csv'.")
                            sys.stdout.flush()

                elif path_type == 'limit':
                    print() # Retained for separation
                    # Get new limit file path
                    while True:
                        new_path = input("Enter the new file path for the limit file (CSV file): ")
                        if new_path.lower().endswith(".csv"):
                            limit_file = new_path
                            break
                        else:
                            print() # Retained for separation
                            print("Invalid path. It must end with '.csv'.")
                            sys.stdout.flush()
                else:
                    print("Invalid option. Please enter 'NB', 'RB', or 'Limit'.")
                    sys.stdout.flush()
                
                print() # Retained for separation
                
                done = input("Are you done making changes? (yes/no): ").strip().lower()
                print() # Retained for separation
                if done == 'yes':
                    break

    return board_file_paths, rb_file_paths, num_boards, num_units, limit_file

def process_dataframes(dataframes):
    """
    Process a list of DataFrames with the following steps:
    - Remove rows with NaN values
    - Drop the 'Test #' column
    - Set 'Description' as the index
    - Transpose the DataFrame
    - Remove the 'Units' index
    - Reset the index
    - Convert all data to float

    Parameters:
    - dataframes: List of DataFrames to process

    Returns:
    - List of processed DataFrames
    """
    processed_dataframes = []

    for df in dataframes:        
        df = df.dropna(thresh=df.shape[1] - 1)  # Remove rows with too many NaNs
        df = df.set_index('Description')  # Use 'Description' as the index
        df = df.drop(df.columns[0], axis=1)  # Drop the first column ('Test #')
        df = df.T  # Transpose the DataFrame
        df = df.drop(index='Units', errors='ignore')  # Remove 'Units' index if it exists
        df = df.reset_index(drop=True)  # Reset the index
        df = df.astype(float)  # Convert all values to float
        
        processed_dataframes.append(df)  # Add the processed DataFrame to the list

    return processed_dataframes

def mean_shift(row):
    """
    Calculate the mean shift percentage based on the provided row data.

    This function computes the mean shift percentage using the 'Delta Mean' value 
    and the specified low and high limits. If either limit is not provided or is 
    NaN, the function returns NaN. If the 'Delta Mean' is zero, the function returns 0.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - 'Delta Mean': The delta mean value.
        - Column indices 2 and 3 for low and high limits, respectively.

    Returns:
    - float: The mean shift percentage rounded to five decimal places, or NaN if limits are invalid.
    """
    # Attempt to convert limits to float and check for NaN in one go
    low_limit = float(row[2]) if len(row) > 2 else np.nan
    high_limit = float(row[3]) if len(row) > 3 else np.nan

    # Return np.nan if either limit is NaN
    if np.isnan(low_limit) or np.isnan(high_limit):
        return np.nan
    
    # Return 0 if Delta Mean is 0
    if row.get('Delta Mean', 0) == 0:
        return 0
    
    # Calculate and return the mean shift percentage
    return np.round(row['Delta Mean'] / (high_limit - low_limit) * 100, 5)

def mean_shift_crit(row):
    """
    Evaluate the mean shift criteria based on the provided row data.

    This function checks the 'Mean Shift' and 'Delta Mean' values against a specified 
    standard deviation limit. It returns "Passed", "Failed", or "For check" based on 
    the evaluation criteria.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - 'Mean Shift': The mean shift value to evaluate.
        - 'Delta Mean': The delta mean value.
        - Column index 4 for the standard deviation limit.

    Returns:
    - str: "Passed", "Failed", or "For check" based on the evaluation criteria.
    """
    mean_shf = row['Mean Shift']
    d_mean = row['Delta Mean']
    sdlot = row[4]

    # Check the conditions
    if np.isnan(mean_shf):  # Check if mean_shf is NaN
        return "Passed" if d_mean <= sdlot else "For check"
    else:
        return "Passed" if mean_shf <= 5 else "Failed"

def sd_ratio(row):
    """
    Calculate the standard deviation ratio based on the provided row data.

    This function computes the ratio of two values from the row. If either of the 
    values is zero, the function returns 0. Otherwise, it returns the ratio rounded 
    to six decimal places.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - Column index 7: The standard deviation for RB at certain unit.
        - Column index 9: The standard deviation for certain NB at certain unit.

    Returns:
    - float: The standard deviation ratio rounded to six decimal places, or 0 if 
      either value is zero.
    """
    if row[7] == 0:
        return 0
    elif row[9] == 0:
        return 0
    else:
        return np.round(row[9] / row[7], 6)

def sd_ratio_crit(row):
    """
    Evaluate the standard deviation ratio criteria based on the provided row data.

    This function checks the 'SD Ratio' value against a threshold of 1.5. 
    It returns "Passed" if the ratio is less than or equal to 1.5, 
    and "For check" otherwise.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - 'SD Ratio': The standard deviation ratio to evaluate.

    Returns:
    - str: "Passed" if the SD Ratio is less than or equal to 1.5, 
           or "For check" if it exceeds 1.5.
    """
    if row["SD Ratio"] <= 1.5:
        return "Passed"
    else:
        return "For check"

def eva_status(row):
    """
    Evaluate the overall status based on mean shift and standard deviation ratio criteria.

    This function checks the evaluation status of 'Mean Shift Criteria' and 
    'SD Ratio Criteria'. It returns "Passed" if both criteria are passed, 
    "Failed" if the mean shift criteria has failed, and "For check" 
    for any other combination of statuses.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - 'Mean Shift Criteria': The evaluation result of the mean shift.
        - 'SD Ratio Criteria': The evaluation result of the standard deviation ratio.

    Returns:
    - str: "Passed" if both criteria are passed, 
           "Failed" if the mean shift criteria has failed, 
           or "For check" for any other combination.
    """
    if row['Mean Shift Criteria'] == "Passed" and row['SD Ratio Criteria'] == "Passed":
        return "Passed"
    elif row['Mean Shift Criteria'] == "Failed":
        return "Failed"
    else:
        return "For check"

def calculate_cp_rb(row):
    """
    Calculate the Cp and RB value based on the provided row data.

    This function computes the Cp and Rb value using the formula:
    (row[3] - row[2]) / (6 * row[7]). It checks for specific conditions 
    before performing the calculation, returning NaN if any of the 
    required values are zero or missing.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[7]: The standard deviation for RB

    Returns:
    - float: The calculated Cp and RB value rounded to two decimal places, 
             or NaN if any of the required values are zero or missing.
    """
    # Check if row[7] is 0
    if row[7] == 0:
        return np.nan
    
    # Check for empty or "NA" values
    if row[2] is np.nan or row[3] is np.nan or row[7] is np.nan:
        return np.nan
    
    # Perform the calculation
    result = (row[3] - row[2]) / (6 * row[7])
    
    # Round the result to 2 decimal places
    return round(result, 2)

def calculate_cpk_rb(row):
    """
    Calculate the CpK and RB value based on the provided row data.

    This function computes the CpK value using the available specification limits 
    and standard deviation. It checks for NaN values and ensures that the standard 
    deviation is not zero before performing the calculation. The function returns 
    the CpK value rounded to two decimal places along with a decision message 
    indicating the capability status.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[6]: The mean for RB at certain unit.
        - row[7]: The standard deviation for RB

    Returns:
    - tuple: A tuple containing:
        - float: The calculated CpK value rounded to two decimal places, 
                 or NaN if the required values are missing or invalid.
        - str: A decision message indicating the capability status ("Not capable" 
               or "Good capable").
    """
    # Check for NaN values and if row[7] is 0
    if np.isnan(row[2]) and np.isnan(row[3]) or row[7] == 0:
        return np.nan, "N/A"
    
    # Calculate cpk based on the available values
    if np.isnan(row[2]):
        cpk = (row[3] - row[6]) / (3 * row[7])
    elif np.isnan(row[3]):
        cpk = (row[6] - row[2]) / (3 * row[7])
    else:
        cpk = min(row[3] - row[6], row[6] - row[2]) / (3 * row[7])

    # Round cpk to 2 decimal places
    cpk = round(cpk, 2)

    # Concise decision statements
    if cpk < 1.3:
        decision_message = "Not capable"
    else:
        decision_message = "Good capable"

    return cpk, decision_message

def calculate_cp_nb(row):
    """
    Calculate the Cp value based on the provided row data.

    This function computes the Cp value using the formula:
    (row[3] - row[2]) / (6 * row[12]). It checks for specific conditions 
    before performing the calculation, returning NaN if any of the 
    required values are zero or missing.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[12]: The standard deviation for certain NB

    Returns:
    - float: The calculated Cp value rounded to two decimal places, 
             or NaN if any of the required values are zero or missing.
    """
    # Check if row[12] is 0
    if row[12] == 0:
        return np.nan
    
    # Check for empty or "NA" values
    if np.isnan(row[2]) or np.isnan(row[3]) or np.isnan(row[12]):
        return np.nan
    
    # Perform the calculation
    result = (row[3] - row[2]) / (6 * row[12])
    
    # Round the result to 2 decimal places
    return round(result, 2)

def calculate_cpk_nb(row):
    """
    Calculate the CpK value based on the provided row data.

    This function computes the CpK value using the available specification limits 
    and standard deviation. It checks for NaN values and ensures that the standard 
    deviation is not zero before performing the calculation. The function returns 
    the CpK value rounded to two decimal places along with a decision message 
    indicating the capability status.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[11]: The mean for certain NB.
        - row[12]: The standard deviation for ceratin NB.

    Returns:
    - tuple: A tuple containing:
        - float: The calculated CpK value rounded to two decimal places, 
                 or NaN if the required values are missing or invalid.
        - str: A decision message indicating the capability status ("Not capable" 
               or "Good capable").
    """
    # Check for NaN values and if row[12] is 0
    if np.isnan(row[2]) and np.isnan(row[3]) or row[12] == 0:
        return np.nan, "N/A"
    
    # Calculate cpk based on the available values
    if np.isnan(row[2]):
        cpk = (row[3] - row[11]) / (3 * row[12])
    elif np.isnan(row[3]):
        cpk = (row[11] - row[2]) / (3 * row[12])
    else:
        cpk = min(row[3] - row[11], row[11] - row[2]) / (3 * row[12])

    # Round cpk to 2 decimal places
    cpk = round(cpk, 2)

    # Concise decision statements
    if cpk < 1.3:
        decision_message = "Not capable"
    else:
        decision_message = "Good capable"

    return cpk, decision_message

def check_value(col):
    """
    Check the value in the specified column and return a corresponding message.

    This function evaluates the value in the fourth element of the provided 
    column (index 3) and returns a message based on its content. 

    Parameters:
    - col: A list or array-like structure where the fourth element (index 3) 
           is evaluated.

    Returns:
    - str: A message indicating the status based on the value:
        - An empty string if the value is an empty string.
        - "Good to release if no concern" if the value is 0.
        - "Not acceptable" for any other value.
    """
    if col[3] == "":
        return ""
    elif col[3] == 0:
        return "Good to release if no concern"
    else:
        return "Not acceptable"

def autosize_columns(worksheet):
    """
    Auto-adjust the width of columns in the given worksheet.

    This function iterates through all rows in the specified worksheet and calculates 
    the maximum length of the content in each column. It then sets the width of each 
    column based on the calculated maximum lengths, adding a small padding for better 
    visibility.

    Parameters:
    - worksheet: An instance of an openpyxl worksheet where the column widths 
                  need to be adjusted.
    
    Returns:
    - None: This function modifies the worksheet in place and does not return a value.
    """
    column_widths = []

    # Iterate through all rows in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if cell is not None:
                cell_length = len(str(cell))  # Get the length of the cell content
                if len(column_widths) > i:
                    if cell_length > column_widths[i]:
                        column_widths[i] = cell_length
                else:
                    column_widths.append(cell_length)

    # Set the width of each column based on the calculated maximum lengths
    for i, column_width in enumerate(column_widths, 1):  # Start at 1 for column indexing
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 0.5  # Adding padding


# ## Display Banner & Asking Input

# Display the banner to the user
display_banner()

# Retrieve product information
product_info = get_product_information()

# Retrieve setup information
setup_info = get_setup_information()

# Obtain data related to files, number of boards, number of units, and limits
nb_file, rb_file, num_boards, num_units, limit = get_data()


# ## Data Processing

# Read the CSV file into a DataFrame
limit = pd.read_csv(limit)

# Reset the index of the DataFrame, dropping the old index
limit.reset_index(drop=True, inplace=True)

# Convert the 'Test #' column to strings and remove '\t'
limit.iloc[:, 0] = limit.iloc[:, 0].astype(str).str.replace(r'\t', '', regex=True)

# Remove rows with too many NaNs
limit.dropna(thresh=limit.shape[1] - 3, inplace=True)  

# Fill any NaN values with numpy's NaN
limit.fillna(np.nan, inplace=True)

# Convert the values in the third column to float
limit.iloc[:, 2] = limit.iloc[:, 2].astype('float')

# Convert the values in the fourth column to float
limit.iloc[:, 3] = limit.iloc[:, 3].astype('float')


# Initialize an empty list to hold the DataFrames
rb = []

# Loop through the unit numbers
for i in range(num_units):
    # Construct the filename based on the unit number
    file_path = rb_file[i]
    
    # Read the CSV file and append the DataFrame to the list
    try:
        df = pd.read_csv(file_path)  # Read the CSV file into a DataFrame
        rb.append(df)  # Append the DataFrame to the list
    except FileNotFoundError:
        print(f"File not found: {file_path}")  # Handle the case where the file is not found
    except Exception as e:
        print(f"An error occurred while reading {file_path}: {e}")  # Handle any other exceptions


# Initialize an empty list to hold the DataFrames for each board
nb = []

# Loop to get input paths for each board and its units
for i in range(num_boards):
    board_units = []  # Initialize a list to hold the DataFrames for the current board
    
    for j in range(num_units):
        # Construct the input path based on board and unit numbers
        file_path = nb_file[i][j]

        # Read the DataFrame from the constructed input path
        try:
            df = pd.read_csv(file_path)  # Read the CSV file into a DataFrame
            board_units.append(df)  # Append the DataFrame for the current unit
            
        except FileNotFoundError:
            print(f"File not found: {file_path}")  # Handle the case where the file is not found
        except Exception as e:
            print(f"An error occurred while reading {file_path}: {e}")  # Handle any other exceptions
    
    # Append the current board's units to the main list
    nb.append(board_units)  # Add the list of DataFrames for the current board to the main list


# Process the list of DataFrames for the 'rb' variable
rb_mod = process_dataframes(rb)

# Initialize an empty list to hold processed DataFrames for each board
nb_mod = []

# Loop through each board in the 'nb' list
for board in nb:
    # Process the DataFrames for the current board
    processed_units = process_dataframes(board)
    
    # Append processed DataFrames for each board to the main list
    nb_mod.append(processed_units)  # Add the processed units for the current board


# ## Correlation Report

# Initialize empty lists to hold the mean and standard deviation DataFrames
mean_rb = []
std_rb = []

# Loop through each modified DataFrame in the 'rb_mod' list
for df_mod in rb_mod:
    # Calculate the mean for each column and format it to six decimal places
    mean = df_mod.mean().apply(lambda x: f'{x:.6f}').astype(float)
    
    # Calculate the standard deviation for each column (using population standard deviation) and format it to five decimal places
    std = df_mod.std(ddof=0).apply(lambda x: f'{x:.5f}').astype(float)
    
    # Append the mean and standard deviation to their respective lists
    mean_rb.append(mean)
    std_rb.append(std)

# Reset the index of each mean DataFrame to create a clean list of DataFrames
mean_rb_clean = [df.reset_index(drop=True) for df in mean_rb]

# Reset the index of each standard deviation DataFrame to create a clean list of DataFrames
std_rb_clean = [df.reset_index(drop=True) for df in std_rb]

# Create a list of column names for the mean DataFrames
columns_mean_rb = [f'Mean RB U{i+1}' for i in range(0, num_units)]

# Create a list of column names for the standard deviation DataFrames
columns_std_rb = [f'SD RB U{i+1}' for i in range(0, num_units)]

# Initialize empty DataFrames to hold the real mean and standard deviation values
realrbmean = pd.DataFrame()
realrbstd = pd.DataFrame()

# Loop through each column name and corresponding mean value
for col, val in zip(columns_mean_rb, mean_rb_clean):
    realrbmean[col] = val  # Assign the mean values to the corresponding column in the DataFrame

# Loop through each column name and corresponding standard deviation value
for col, val in zip(columns_std_rb, std_rb_clean):
    realrbstd[col] = val  # Assign the standard deviation values to the corresponding column in the DataFrame


# Assuming nb_mod is a list of lists of NumPy arrays instead of DataFrames
mean_nb_clean = []  # Initialize the mean list outside the loop
std_nb_clean = []   # Initialize the standard deviation list outside the loop

# Iterate through each inner list in nb_mod
for inner_list in nb_mod:
    mean_row = []  # Temporary list to hold means for the current inner list
    std_row = []   # Temporary list to hold standard deviations for the current inner list
    
    # Iterate through each NumPy array in the inner list
    for arr_mod in inner_list:
        # Calculate mean and standard deviation using NumPy
        mean = np.mean(arr_mod, axis=0)  # Calculate mean along the specified axis
        std = np.std(arr_mod, axis=0, ddof=0)  # Calculate standard deviation along the specified axis
        
        # Append results to respective temporary lists
        mean_row.append(np.round(mean, 6).tolist())  # Round the mean and convert to list
        std_row.append(np.round(std, 5).tolist())    # Round the standard deviation and convert to list
    
    # Append the temporary lists to the main 2D lists
    mean_nb_clean.append(mean_row)  # Add the mean row to the main list
    std_nb_clean.append(std_row)     # Add the std row to the main list

# mean_nb_clean and std_nb_clean are now 2D lists of means and standard deviations

# Create 2D lists to hold column names for means and standard deviations
columns_mean_nb = []
columns_std_nb = []

# Fill the lists with column names for each board
for i in range(num_boards + 1):
    # Generate mean column names for the current board
    mean_columns = [f'Mean NB{i+1} U{j+1}' for j in range(0, num_units)]
    
    # Generate standard deviation column names for the current board
    std_columns = [f'SD NB{i+1} U{j+1}' for j in range(0, num_units)]
    
    # Append the mean columns for the current board to the list
    columns_mean_nb.append(mean_columns)
    
    # Append the standard deviation columns for the current board to the list
    columns_std_nb.append(std_columns)

# Create a list to hold the empty DataFrames for means
realnbmean = []

# Create a list to hold the empty DataFrames for standard deviations
realnbstd = []

# Loop to create empty DataFrames and append them to the realnbmean list
for _ in range(num_boards):
    realnbmean.append(pd.DataFrame())  # Append an empty DataFrame for each board

# Loop to create empty DataFrames and append them to the realnbstd list
for _ in range(num_boards):
    realnbstd.append(pd.DataFrame())  # Append an empty DataFrame for each board

# Populate the realnbmean DataFrames with mean values
for i in range(num_boards):
    for col, val in zip(columns_mean_nb[i], mean_nb_clean[i]):
        realnbmean[i][col] = val  # Assign mean values to the corresponding columns

# Populate the realnbstd DataFrames with standard deviation values
for i in range(num_boards):
    for col, val in zip(columns_std_nb[i], std_nb_clean[i]):
        realnbstd[i][col] = val  # Assign standard deviation values to the corresponding columns

# Initialize a list to hold the calculated limits for each unit
rb_lim_calc = []  # i = unit

# Loop through each unit to calculate limits
for i in range(num_units):
    # Concatenate the limit DataFrame with the mean and standard deviation for the current unit
    temp = pd.concat([limit, realrbmean.iloc[:, i], realrbstd.iloc[:, i]], axis=1)
    
    # Append the concatenated DataFrame to the rb_lim_calc list
    rb_lim_calc.append(temp)

# Create a 2D list to hold the concatenated DataFrames for each board and unit
rb_nbu = []

# Loop through the boards to create the 2D list
for i in range(num_boards):
    temp_list = []  # Temporary list to hold DataFrames for the current board
    for j in range(num_units):
        # Concatenate the DataFrames for the current unit, including limits, means, and standard deviations
        temp = pd.concat([rb_lim_calc[j], realnbmean[i].iloc[:, j], realnbstd[i].iloc[:, j]], axis=1)
        
        # Append the concatenated DataFrame to the temporary list
        temp_list.append(temp)
    
    # Append the temporary list of DataFrames for the current board to the main 2D list
    rb_nbu.append(temp_list)

# Loop through each board
for i in range(num_boards):
    # Loop through each unit within the current board
    for j in range(num_units):
        # Calculate the absolute difference between two columns and round it to 6 decimal places
        rb_nbu[i][j]['Delta Mean'] = rb_nbu[i][j].apply(lambda row: np.round(abs(row[8] - row[6]), 6), axis=1)
        
        # Apply the mean_shift function to calculate the mean shift for each row
        rb_nbu[i][j]['Mean Shift'] = rb_nbu[i][j].apply(mean_shift, axis=1)
        
        # Apply the mean_shift_crit function to evaluate the mean shift criteria for each row
        rb_nbu[i][j]['Mean Shift Criteria'] = rb_nbu[i][j].apply(mean_shift_crit, axis=1)
        
        # Apply the sd_ratio function to calculate the standard deviation ratio for each row
        rb_nbu[i][j]['SD Ratio'] = rb_nbu[i][j].apply(sd_ratio, axis=1)
        
        # Apply the sd_ratio_crit function to evaluate the standard deviation ratio criteria for each row
        rb_nbu[i][j]['SD Ratio Criteria'] = rb_nbu[i][j].apply(sd_ratio_crit, axis=1)
        
        # Apply the eva_status function to determine the result for each unit based on the evaluations
        rb_nbu[i][j]['Result Unit'] = rb_nbu[i][j].apply(eva_status, axis=1)


# Create an empty list to hold the DataFrames for each board
nb_results = [pd.DataFrame() for _ in range(num_boards)]

# Loop through each board to populate the results
for i in range(num_boards):
    # Create a new column for each unit's result within the current board
    for j in range(num_units):
        # Assign the 'Result Unit' data from rb_nbu to the new DataFrame for the current board and unit
        nb_results[i][f'Result NB{i+1} U{j+1}'] = rb_nbu[i][j]['Result Unit']

    # Check if all results for the board are "Passed" and create a new column for the overall result
    all_passed = nb_results[i].eq("Passed").all(axis=1)
    
    # Assign the overall result based on whether all units passed or not
    nb_results[i][f"NB{i+1} Result"] = all_passed.replace({True: 'Passed', False: 'For check'})

# Initialize an empty list to hold the concatenated results
nbtrueresult = []

# Loop through each board to concatenate the limit with the results
for i in range(num_boards):
    # Concatenate the limit DataFrame/Series with the results DataFrame for the current board
    temp = pd.concat([limit, nb_results[i]], axis=1)  # Wrap the arguments in a list
    
    # Append the concatenated DataFrame to the nbtrueresult list
    nbtrueresult.append(temp)


# ## Cpk Report

# Concatenate all DataFrames in rb_mod for each unit into a single DataFrame, ignoring the index
rbdf = pd.concat([rb_mod[i] for i in range(num_units)], ignore_index=True)

# Initialize an empty list to hold the concatenated DataFrames for each board
nbdf = []

# Loop through each board to concatenate the results for all units
for i in range(0, num_boards):
    # Concatenate all DataFrames in nb_mod for the current board across all units, ignoring the index
    temp = pd.concat([nb_mod[i][j] for j in range(num_units)], ignore_index=True)
    
    # Append the concatenated DataFrame for the current board to the nbdf list
    nbdf.append(temp)

# Calculate the mean of the rbdf DataFrame, format it to four decimal places, and convert it to float
meanrbdf = rbdf.mean().apply(lambda x: f'{x:.4f}').astype(float).reset_index(drop=True)

# Initialize an empty list to hold the mean results for each board
meannbdf = []

# Loop through each board to calculate the mean of the corresponding DataFrame
for i in range(0, num_boards):
    # Calculate the mean of the current board's DataFrame, format it to four decimal places, and convert it to float
    temp = nbdf[i].mean().apply(lambda x: f'{x:.4f}').astype(float)
    
    # Append the mean result for the current board to the meannbdf list
    meannbdf.append(temp)

# Reset the index for each DataFrame in the meannbdf list
meannbdf = [df.reset_index(drop=True) for df in meannbdf]

# Calculate the standard deviation of the rbdf DataFrame, format it to four decimal places, and convert it to float
stdrbdf = rbdf.std(ddof=0).apply(lambda x: f'{x:.4f}').astype(float).reset_index(drop=True)

# Initialize an empty list to hold the standard deviation results for each board
stdnbdf = []

# Loop through each board to calculate the standard deviation of the corresponding DataFrame
for i in range(0, num_boards):
    # Calculate the standard deviation of the current board's DataFrame, format it to four decimal places, and convert it to float
    temp = nbdf[i].std(ddof=0).apply(lambda x: f'{x:.4f}').astype(float)
    
    # Append the standard deviation result for the current board to the stdnbdf list
    stdnbdf.append(temp)

# Reset the index for each DataFrame in the stdnbdf list
stdnbdf = [df.reset_index(drop=True) for df in stdnbdf]

# Concatenate the mean and standard deviation DataFrames for rbdf along the columns, ignoring the index
meanstdrbdf = pd.concat([meanrbdf, stdrbdf], axis=1, ignore_index=True)

# Set the column names for the concatenated DataFrame
meanstdrbdf.columns = ['Mean RB', 'SD RB']

# Initialize a list to hold the concatenated mean and standard deviation DataFrames for each board
meanstdnbdf = [pd.DataFrame() for _ in range(num_boards)]

# Loop through each board to concatenate the mean and standard deviation DataFrames
for i in range(0, num_boards):
    # Concatenate the mean and standard deviation DataFrames for the current board along the columns, ignoring the index
    meanstdnbdf[i] = pd.concat([meannbdf[i], stdnbdf[i]], axis=1, ignore_index=True)
    
    # Set the column names for the concatenated DataFrame of the current board
    meanstdnbdf[i].columns = [f'Mean NB{i+1}', f'SD NB{i+1}']


# Concatenate the limit DataFrame with the mean and standard deviation DataFrame for rbdf along the columns
rbcpcpk = pd.concat([limit, meanstdrbdf], axis=1)

# Apply the calculate_cp_rb function to each row of the concatenated DataFrame to calculate the Cp value for rb
rbcpcpk["Cp RB"] = rbcpcpk.apply(calculate_cp_rb, axis=1)

# Apply the calculate_cpk_rb function to each row of the concatenated DataFrame to calculate the Cpk values
# The result is expanded into two new columns: 'Cpk RB' and 'Cpk RB Result'
rbcpcpk[['Cpk RB', 'Cpk RB Result']] = rbcpcpk.apply(calculate_cpk_rb, axis=1, result_type='expand')


# Initialize an empty list to hold the DataFrames for each board's Cp and Cpk results
nbcpkresult = []

# Loop through each board to concatenate the rbcpcpk DataFrame with the corresponding mean and standard deviation DataFrame
for i in range(0, num_boards):
    # Concatenate the rbcpcpk DataFrame with the mean and standard deviation DataFrame for the current board
    temp = pd.concat([rbcpcpk, meanstdnbdf[i]], axis=1, ignore_index=False)
    
    # Append the concatenated DataFrame to the nbcpkresult list
    nbcpkresult.append(temp)

# Loop through each board to calculate the Cp values and add them to the corresponding DataFrame
for i in range(0, num_boards):
    # Apply the calculate_cp_nb function to each row of the current board's DataFrame to calculate the Cp value
    nbcpkresult[i][f"Cp NB{i+1}"] = nbcpkresult[i].apply(calculate_cp_nb, axis=1)

# Loop through each board to calculate the Cpk values and add them to the corresponding DataFrame
for i in range(0, num_boards):
    # Apply the calculate_cpk_nb function to each row of the current board's DataFrame to calculate the Cpk values
    # The result is expanded into two new columns: 'Cpk NB' and 'Cpk NB Result'
    nbcpkresult[i][[f'Cpk NB{i+1}', f'Cpk NB{i+1} Result']] = nbcpkresult[i].apply(calculate_cpk_nb, axis=1, result_type='expand')


# ## Correlation Result Summary Table

# Initialize an empty DataFrame to hold the correlation table
corrtable = pd.DataFrame()

# Creating corrtable and adding necessary columns
# Add a column for the test card names, formatted with the product info and board index
corrtable['Test Card'] = [f"{product_info['Test Card Name']}_NB{i+1}" for i in range(num_boards)]

# Add a column for the count of passed tests across all units for each board
corrtable['Passed test all units'] = [nb_results[i][f"NB{i+1} Result"].str.contains('Passed').sum() for i in range(num_boards)]

# Add a column for the count of tests marked 'For check' across all units for each board
corrtable['For Check test all units'] = [nb_results[i][f"NB{i+1} Result"].str.contains('For check').sum() for i in range(num_boards)]

# Add a column for the count of failed tests across all units for each board
corrtable['Failed test all units'] = [nb_results[i][f"NB{i+1} Result"].str.contains('Failed').sum() for i in range(num_boards)]

# Adding unit-specific columns for each unit
for j in range(num_units):
    # Add a column for the count of passed tests for the current unit across all boards
    corrtable[f'Passed test U{j+1}'] = [nb_results[i][f"Result NB{i+1} U{j+1}"].str.contains('Passed').sum() for i in range(num_boards)]
    
    # Add a column for the count of tests marked 'For check' for the current unit across all boards
    corrtable[f'For Check test U{j+1}'] = [nb_results[i][f"Result NB{i+1} U{j+1}"].str.contains('For check').sum() for i in range(num_boards)]
    
    # Add a column for the count of failed tests for the current unit across all boards
    corrtable[f'Failed test U{j+1}'] = [nb_results[i][f"Result NB{i+1} U{j+1}"].str.contains('Failed').sum() for i in range(num_boards)]

# Add a column for the total number of tests conducted (assuming all boards have the same number of results)
corrtable['Total test'] = len(nb_results[0])

# Apply a function to check values and add remarks to the DataFrame
corrtable['Remarks'] = corrtable.apply(check_value, axis=1)

# Set the 'Test Card' column as the index of the DataFrame
corrtable.set_index('Test Card', inplace=True)

# Transpose the DataFrame to switch rows and columns
corrtable = corrtable.T

# Optional: Rename the axes for clarity
corrtable.rename_axis("Test Card", axis=0, inplace=True)
corrtable.rename_axis("Index", axis=1, inplace=True)

# Reset the index to convert the index back into a column
corrtable.reset_index(inplace=True)

# Create a DataFrame from the product_info dictionary, using the index as the first column and 'Details' as the second column
df_product_info = pd.DataFrame.from_dict(product_info, orient='index', columns=['Details']).reset_index()

# Create a DataFrame from the setup_info dictionary, using the index as the first column and 'Details' as the second column
df_setup_info = pd.DataFrame.from_dict(setup_info, orient='index', columns=['Details']).reset_index()

# Rename the columns of the product info DataFrame for clarity
df_product_info.columns = ['Product Info', 'Details']

# Rename the columns of the setup info DataFrame for clarity
df_setup_info.columns = ['Setup Info', 'Details']


# ## Generate output file

# Define the output file name for the Excel report, incorporating the test card name and the current date

current_date = datetime.now().strftime("%d-%m-%Y")  # Get the current date in YYYY-MM-DD format

output_file = f'{product_info["Test Card Name"]}_Correlation_Report_{current_date}.xlsx'

# Create an Excel writer object to write multiple DataFrames to an Excel file
with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
    # Write additional DataFrames to specific sheets first
    df_product_info.to_excel(writer, sheet_name='Info', index=False)  # Write product info to 'Info' sheet
    df_setup_info.to_excel(writer, sheet_name='Info', index=False, startrow=len(df_product_info) + 2, header=True)  # Write setup info below product info
    corrtable.to_excel(writer, sheet_name='Correlation Summary', index=False)  # Write correlation summary to its own sheet

    # Write the first set of DataFrames (rb_nbu) to separate sheets
    for row_index, row in enumerate(rb_nbu):
        for col_index, df in enumerate(row):
            sheet_name = f'NB{row_index + 1} U{col_index + 1} Corr'  # Create a sheet name based on the row and column index
            df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write the DataFrame to the specified sheet
            worksheet = writer.sheets[sheet_name]  # Access the worksheet for further formatting

    # Apply conditional formatting for specific criteria in the worksheets
    for row_index, row in enumerate(rb_nbu):
        for col_index, df in enumerate(row):
            sheet_name = f'NB{row_index + 1} U{col_index + 1} Corr'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            # Apply PatternFill for `Mean Shift Criteria` and `SD Ratio Criteria`
            for idx, cell in enumerate(worksheet['M'][1:], start=1):  # Column M: Mean Shift Criteria
                value = cell.value
                if value == "Passed":
                    cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
                elif value == "For check":
                    cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
                elif value == "Failed":
                    cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")

            for idx, cell in enumerate(worksheet['O'][1:], start=1):  # Column O: SD Ratio Criteria
                value = cell.value
                if value == "Passed":
                    cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
                elif value == "For check":
                    cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
                elif value == "Failed":
                    cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")

            for idx, cell in enumerate(worksheet['P'][1:], start=1):  # Column P: SD Ratio Criteria
                value = cell.value
                if value == "Passed":
                    cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
                elif value == "For check":
                    cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
                elif value == "Failed":
                    cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")

    # Write the third set of DataFrames (nbtrueresult) to separate sheets
    for i, df in enumerate(nbtrueresult):
        sheet_name = f'NB{i + 1} Results'  # Create a sheet name for the results
        df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write the DataFrame to the specified sheet
        worksheet = writer.sheets[sheet_name]  # Access the worksheet for further formatting

        # Apply PatternFill for `Result NB# U#`
        for col_index in range(1, len(df.columns)):  # Dynamic column range
            if col_index < len(df.columns) - 1:  # Exclude the last "NB# Result" column
                column_letter = get_column_letter(col_index + 1)
                for cell in worksheet[column_letter][1:]:  # Exclude header row
                    value = cell.value
                    if value == "Passed":
                        cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
                    elif value == "For check":
                        cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
                    elif value == "Failed":
                        cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")

        # Apply PatternFill for `NB# Result` (last column)
        result_column_letter = get_column_letter(len(df.columns))  # Last column for "NB# Result"
        for cell in worksheet[result_column_letter][1:]:  # Exclude header row
            value = cell.value
            if value == "Passed":
                cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
            elif value == "For check":
                cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")

    # Write the second set of DataFrames (nbcpkresult) to separate sheets
    for i, df in enumerate(nbcpkresult):
        sheet_name = f'NB{i + 1} CPK'  # Create a sheet name for CPK results
        df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write the DataFrame to the specified sheet
        worksheet = writer.sheets[sheet_name]  # Access the worksheet for further formatting

        # Apply PatternFill for `Cpk RB Result` and `Cpk NB# Result`
        for idx, cell in enumerate(worksheet['K'][1:], start=1):  # Column K
            value = cell.value
            if value == "Good capable":
                cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
            elif value == "Not capable":
                cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")
            elif value == "N/A":
                cell.fill = PatternFill(start_color="95afc0", end_color="95afc0", fill_type="solid")

        for idx, cell in enumerate(worksheet['P'][1:], start=1):  # Column P
            value = cell.value
            if value == "Good capable":
                cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
            elif value == "Not capable":
                cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")
            elif value == "N/A":
                cell.fill = PatternFill(start_color="95afc0", end_color="95afc0", fill_type="solid")

    # Access the workbook and the writer's worksheets
    workbook = writer.book

    # Define a border style for the cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    fill_color = PatternFill(start_color='82ccdd', end_color='82ccdd', fill_type='solid')
    fill_color_info = PatternFill(start_color='74b9ff', end_color='74b9ff', fill_type='solid')
    fill_color_corr = PatternFill(None)

    # Adjust column widths, add autofilters, and apply formatting
    for sheet_name in writer.sheets:
        worksheet = workbook[sheet_name]

        # Adjust column widths for the current sheet
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            for cell in worksheet[1]:  # Access the first row
                cell.fill = fill_color  # Apply the fill to each cell in the first row

            for cell in column:
                cell.border = thin_border  # Apply border to each cell in the column

        # Add autofilter to all sheets except 'Info'
        if sheet_name != 'Info':
            worksheet.auto_filter.ref = worksheet.dimensions

        # Apply borders and formatting to specific sheets
        if sheet_name == 'Info':
            for cell in worksheet[1]:  # Access the first row
                cell.fill = fill_color_info  # Apply the fill to each cell in the first row

            for cell in worksheet[8]:  # Access the eighth row
                cell.fill = fill_color_info  # Apply the fill to each cell in the eighth row

            # Apply borders to the first table (df_product_info)
            for row in worksheet.iter_rows(min_row=1, max_row=len(df_product_info) + 1, min_col=1, max_col=len(df_product_info.columns)):
                for cell in row:
                    cell.border = thin_border  # Apply border to each cell in the first table

            # Apply borders to the second table (df_setup_info)
            for row in worksheet.iter_rows(min_row=len(df_product_info) + 3, max_row=len(df_product_info) + len(df_setup_info) + 3, min_col=1, max_col=len(df_setup_info.columns)):
                for cell in row:
                    cell.border = thin_border  # Apply border to each cell in the second table

            for cell in worksheet[7]:  # Access the eighth row
                cell.border = None  # Remove border from the eighth row

        if sheet_name == 'Correlation Summary':
            for row in worksheet.iter_rows(min_row=1, max_row=len(corrtable) + 1, min_col=1, max_col=len(corrtable.columns)):
                for cell in row:
                    cell.border = thin_border  # Apply border to each cell in the correlation summary
                    cell.fill = fill_color_corr  # Apply fill color to each cell

                    # Align all cells to the left
                    cell.alignment = Alignment(horizontal='left')

                    # Make the first column bold
                    if cell.column == 1:  # Check if it's the first column
                        cell.font = Font(bold=True)

        # Freeze the first row for all sheets except 'Info' and 'Correlation Summary'
        if sheet_name not in ['Info', 'Correlation Summary']:
            worksheet.freeze_panes = worksheet['A2']  # Freeze the first row

        # Align the first column to the left for all sheets except 'Info' and 'Correlation Summary'
        if sheet_name not in ['Info', 'Correlation Summary']:
            for cell in worksheet['A'][1:]:  # Access the first column (A), excluding the header
                cell.alignment = Alignment(horizontal='left')  # Set alignment to left

# Print a message indicating that the DataFrames have been written to the specified output file
print()
print(f"Report have been written to '{output_file}'")


# Set the timer duration in seconds
timer_duration = 10  # Change this to your desired duration

# Call the thank_you function to display a message or perform an action
thank_you()

# Countdown loop to display the remaining time
for remaining in range(timer_duration, 0, -1):
    # Print the remaining time, overwriting the same line in the terminal
    print(f"The terminal will close in {remaining} seconds...", end='\r')  # Use '\r' to overwrite the line
    time.sleep(1)  # Pause execution for 1 second

# Exit the program after the countdown is complete
sys.exit()
